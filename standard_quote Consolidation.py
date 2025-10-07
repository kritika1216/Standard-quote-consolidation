import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

# -- Change to your folder path as needed --
input_folder = "consolidatedmfgportfoliofiles65"
output_file = "Manufacturers_consolidated.xlsx"

output_headers = [
    "BRAND NAME", "MFG NAME", "MFG_THERAPY NAME", "COMPOSITION", "UPP", "UOM",
    "UOM(DATA TEAM)", "UPP (DATA TEAM)", "HSN CODE", "MRP (RS.)",
    "MRP PER UNIT (DATA TEAM)", "RATE WO GST", "COST WO GST PER UNIT LEVEL (DATA TEAM)",
    "GST", "DIVISION", "GENERIC", "FORM", "DOSE", "MANUFACTURER_NAME",
    "BRAND", "UOM.1", "UNIT_PER_PACK", "GROUP_NAME", "SUBCATEGORY_NAME",
    "CATEGORY_NAME", "BOTTLE TYPE", "PACKING", "REMARKS",
]

# Header color mapping
color_map = {
    "UOM(DATA TEAM)":           "FFFF00",
    "UPP (DATA TEAM)":          "FFFF00",
    "MRP PER UNIT(DATA TEAM)":  "FFFF00",
    "COST WO GST PER UNIT LEVEL(DATA TEAM)": "FFFF00",
}
red_start = "GENERIC"
red_color = "FF0000"
found_red = False
header_colors = []
for h in output_headers:
    if h in color_map:
        header_colors.append(color_map[h])
    elif h == red_start or found_red:
        header_colors.append(red_color)
        found_red = True
    else:
        header_colors.append("8DB4E2")

def normalize(col):
    return str(col).upper().replace(" ", "").replace("_", "")

def find_data_start_row(df, output_headers):
    norm_headers = [normalize(h) for h in output_headers]
    for idx in range(len(df)):
        row = df.iloc[idx]
        count = sum(normalize(c) in norm_headers for c in row)
        if count >= len(output_headers) // 2:
            return idx
    return None

def normalize_gst_value(val):
    if pd.isnull(val):
        return ""
    try:
        if isinstance(val, str):
            val = val.strip().replace("%", "")
            if val == "":
                return ""
            num = float(val)
        else:
            num = float(val)
        if 0 <= num <= 1:
            num = num * 100
        num_rounded = round(num, 2)
        if num_rounded.is_integer():
            return int(num_rounded)
        else:
            return num_rounded
    except:
        s = str(val).replace("%", "").strip()
        return s if s else ""

def round_to_two_decimals(val):
    """Round ONLY if val is a float or int; preserve text as-is."""
    if isinstance(val, (float, int)) and not isinstance(val, bool):
        return round(val, 2)
    # Try to coerce string that is numeric as well, but skip empty text
    try:
        if isinstance(val, str) and val.strip() and not any(c.isalpha() for c in val):
            num = float(val)
            return round(num, 2)
    except:
        pass
    return val

def round_numeric_columns_mix(df, exclude_columns=None):
    """For each column, round numbers to 2dp; text is preserved."""
    if exclude_columns is None:
        exclude_columns = []
    for col in df.columns:
        if col in exclude_columns:
            continue
        df[col] = df[col].apply(round_to_two_decimals)
        df[col] = df[col].replace({np.nan: ""})
    return df

all_dataframes = []

for filename in os.listdir(input_folder):
    if filename.lower().endswith(('.xlsx', '.xls')):
        file_path = os.path.join(input_folder, filename)
        try:
            xl = pd.ExcelFile(file_path)
            found_data = False
            for sheet in xl.sheet_names:
                tmp_df = xl.parse(sheet, header=None)
                data_start_row = find_data_start_row(tmp_df, output_headers)
                if data_start_row is not None:
                    df = pd.read_excel(file_path, sheet_name=sheet, header=data_start_row)
                    df_columns_norm = [normalize(c) for c in df.columns]
                    col_map = {}
                    for idx, h in enumerate(output_headers):
                        nh = normalize(h)
                        if nh in df_columns_norm:
                            col_map[h] = df.columns[df_columns_norm.index(nh)]
                        else:
                            col_map[h] = None
                    result_df = pd.DataFrame()
                    for h in output_headers:
                        if col_map[h] is not None:
                            # Don't force to str! Keep dtype as loaded.
                            result_df[h] = df[col_map[h]].replace(["nan", "NaN", "None", "NaT"], "", regex=True)
                        else:
                            result_df[h] = ""
                    # Special normalize GST column
                    result_df["GST"] = result_df["GST"].apply(normalize_gst_value)
                    # List of non-rounded columns
                    non_rounded = [
                        "GST", "BRAND NAME", "MFG NAME", "MFG_THERAPY NAME",
                        "COMPOSITION", "UOM", "UOM(DATA TEAM)", "HSN CODE",
                        "DIVISION", "GENERIC", "FORM", "DOSE", "MANUFACTURER_NAME",
                        "BRAND", "UOM.1", "GROUP_NAME", "SUBCATEGORY_NAME",
                        "CATEGORY_NAME", "BOTTLE TYPE", "PACKING", "REMARKS"
                    ]
                    result_df = round_numeric_columns_mix(result_df, exclude_columns=non_rounded)
                    all_dataframes.append(result_df)
                    found_data = True
                    break
            if not found_data:
                blank_df = pd.DataFrame(columns=output_headers)
                all_dataframes.append(blank_df)
        except Exception as e:
            print(f"Error in {filename}: {e}")
            blank_df = pd.DataFrame(columns=output_headers)
            all_dataframes.append(blank_df)

if all_dataframes:
    combined_df = pd.concat(all_dataframes, ignore_index=True)
else:
    combined_df = pd.DataFrame(columns=output_headers)

combined_df.to_excel(output_file, index=False)

# Apply header colors using openpyxl
wb = load_workbook(output_file)
ws = wb.active
for idx, cell in enumerate(ws[1], 0):
    fill = PatternFill(start_color=header_colors[idx], end_color=header_colors[idx], fill_type="solid")
    cell.fill = fill
wb.save(output_file)

print(f"Data consolidated and formatted (mixed text/numbers, decimals max 2dp) in '{output_file}'.")
